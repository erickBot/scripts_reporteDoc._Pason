import openpyxl
from datetime import datetime, date, time, timedelta
import calendar
import pandas as pd
import numpy as np
import xlsxwriter

list_sheetbyName = []
list_diasVigencia = []
list_contador = []

nameSheet = []
dateExpiracion = []
Item = []
diasPorExpirar = []

def run():
    j = 0
    #abre archivo con toda la informacion
    name_file = "Control_Documentos.xlsx"
    file = openpyxl.load_workbook(name_file)
    #<file.sheetnames> devuelve los nombres de las hojas en una lista
    list_sheetbyName = file.sheetnames
    #crea archivo nuevo para el reporte, y lo cierra
    wb = openpyxl.Workbook()
    ws = wb.create_sheet('resumen', 0)
    ws.sheet_properties.tabColor = 'FFC0CB'
    wb.save('Reporte_Documentos_Por_Vencer.xls')

    #itera hoja por hoja
    for name in list_sheetbyName:
        sheet = file.get_sheet_by_name(name)
        #devuelve el numero maximo de filas por hoja (sheet)
        num_row = sheet.max_row
        cell_total = sheet['D4':'D' + str(num_row)]
        #devuelve la fecha actual del computador
        dateNow = datetime.now()
        lookCell(sheet, cell_total, dateNow, name)
        #salta a las funciones para llenar las celdas con datos nuevos
        llenar_cell(sheet, 'F4', 'F' + str(num_row))
        llenar_cell_Estado(sheet, 'E4', 'E' + str(num_row) )
        #vaciar listas
        list_diasVigencia.clear()
        list_contador.clear()
        #generar reporte por cada iteracion de hojas
        generarReporte()
        #vaciar listas
        nameSheet.clear()  #nombre de hoja
        Item.clear()       #nombre de documento
        dateExpiracion.clear()      #fecha de expiracion
        diasPorExpirar.clear()      #dias de vigencia o expirados
    
    #guarda archivo
    file.save(name_file)

def generarReporte():
    reporte = openpyxl.load_workbook('Reporte_Documentos_Por_Vencer.xlsx')
    hoja = reporte.get_sheet_by_name('resumen')
    numRows = hoja.max_row
    print(numRows)
    #escribe en hoja de reporte si es que hay datos en la lista nameSheet
    for name in nameSheet:
        if (numRows > 1):
            numRows = numRows + 1

        hoja['A' + str(numRows + 1)] = name + ':'
        hoja['A' + str(numRows + 2)] = 'Item'
        hoja['B' + str(numRows + 2)] = 'Fecha Expiracion'
        hoja['C' + str(numRows + 2)] = 'Dias Vigencia'
        for y, v in enumerate(Item):
            hoja['A' + str(numRows + 3 + y)] = v
        for y, v in enumerate(dateExpiracion):
            hoja['B' + str(numRows + 3 + y)] = v
        for y, v in enumerate(diasPorExpirar):
            hoja['C' + str(numRows + 3 + y)] = v

        reporte.save('Reporte_Documentos_Por_Vencer.xlsx')


def lookCell(sheet, cell_total, dateNow, name):
    expirados = 0
    porVencer = 0
    vigentes = 0
    i = 0

    for row in cell_total:
        for cell in row:
            dateCell= cell.value
            dias = dateCell - dateNow
            dias = dias.days
            list_diasVigencia.append(dias)

            if dias < 0:
                expirados += 1
                dateExpiracion.append(cell.value)
                diasPorExpirar.append(dias)
                Item.append(sheet['B' + str(i + 4)].value)
            if dias > 0 and dias < 30:
                porVencer += 1
                dateExpiracion.append(cell.value)
                diasPorExpirar.append(dias)
                Item.append(sheet['B' + str(i + 4)].value)
            if dias > 30:
                vigentes += 1
            
            i += 1
    
    if (porVencer >0 or expirados > 0):
        nameSheet.append(name)


    list_contador.append(porVencer)
    list_contador.append(expirados)
    list_contador.append(vigentes)


def llenar_cell(sheet, cellInit, cellEnd):
    i = 0
    cell_list = sheet[cellInit : cellEnd]
    for row in cell_list:
        for cell in row:
            cell.value = list_diasVigencia[i]
            i += 1

    sheet['L3'] = list_contador[0] #porVencer
    sheet['L4'] = list_contador[1] #expirados
    sheet['L5'] = list_contador[2] #vigentes

    #print(list_diasVigencia)

def llenar_cell_Estado(sheet, cellInit, cellEnd):
    i = 0
    cell_list = sheet[cellInit : cellEnd]
    for row in cell_list:
        for cell in row:
            if list_diasVigencia[i] > 0:
                cell.value = "VIGENTE"
            if list_diasVigencia[i] < 0:
                cell.value = "EXPIRADO"
            i += 1

if __name__ == "__main__":
    run()
